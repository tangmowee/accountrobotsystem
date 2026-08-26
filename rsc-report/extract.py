#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ดึงข้อมูลจากไฟล์ Excel + รายงานเดิม -> data/payload.json สำหรับส่วนใหม่ของรายงาน"""
import openpyxl, json, re, datetime as dt
from collections import defaultdict

M = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul']
out = {}

# ---------------------------------------------------------------- 1) KPI ทางการ (Jan–Jul)
wb = openpyxl.load_workbook('data/KPI_PU_JanJul_2026.xlsx', data_only=True)
ws = wb['KPI-(ข้อ1)']
num = lambda v: float(v) if isinstance(v, (int, float)) else None
out['otd'] = {
    'target': ws.cell(7, 8).value * 100,                       # 90%
    'po':     [num(ws.cell(9, c).value) for c in range(3, 10)],
    'ontime': [num(ws.cell(10, c).value) for c in range(3, 10)],
    'pct':    [num(ws.cell(11, c).value) for c in range(3, 10)],
    'totPo': num(ws.cell(9, 15).value), 'totOn': num(ws.cell(10, 15).value),
    'totPct': num(ws.cell(11, 15).value),
}
ws2 = wb['KPI-(ข้อ2)']
out['csKpi'] = {
    'target': ws2.cell(7, 8).value * 100,                      # 5%
    'cur':  [num(ws2.cell(9, c).value) for c in range(3, 10)],
    'new':  [num(ws2.cell(10, c).value) for c in range(3, 10)],
    'save': [num(ws2.cell(11, c).value) for c in range(3, 10)],
    'pct':  [num(ws2.cell(12, c).value) for c in range(3, 10)],
    'totCur': num(ws2.cell(9, 15).value), 'totSave': num(ws2.cell(11, 15).value),
    'totPct': num(ws2.cell(12, 15).value),
}

# ---------------------------------------------------------------- 2) รายการรับของ Jun / Jul
DELF = {'Jun': ('data/Data_Delivery_June.xlsx', 'Data Delivery June', 8),
        'Jul': ('data/Data_Delivery_Jul_2026.xlsx', 'PO-KPI-Jul', 1)}
rows = []
for m, (f, sn, hr) in DELF.items():
    w = openpyxl.load_workbook(f, data_only=True); s = w[sn]
    hdr = {s.cell(hr, c).value: c for c in range(1, s.max_column + 1) if s.cell(hr, c).value}
    for r in range(hr + 1, s.max_row + 1):
        no = s.cell(r, hdr['No.']).value
        if not isinstance(no, int):
            continue
        g = lambda k: s.cell(r, hdr[k]).value if k in hdr else None
        rows.append(dict(m=m, sup=(g('Suppliers-Name') or '').strip(),
                         det=(g('Details') or '').strip(), tot=float(g('Totals Price') or 0),
                         po=g('PO NO.'), ab=g('Ability'), diff=g('Diff')))

def bucket(d):
    if d >= 0: return 'ตรงกำหนด (Ontime)'
    d = -d
    return ('ล่าช้า 1–3 วัน' if d <= 3 else 'ล่าช้า 4–7 วัน' if d <= 7
            else 'ล่าช้า 8–14 วัน' if d <= 14 else 'ล่าช้ามากกว่า 14 วัน')

BK = ['ตรงกำหนด (Ontime)', 'ล่าช้า 1–3 วัน', 'ล่าช้า 4–7 วัน', 'ล่าช้า 8–14 วัน', 'ล่าช้ามากกว่า 14 วัน']
out['buckets'] = {}
for m in ['Jun', 'Jul']:
    rs = [x for x in rows if x['m'] == m]
    b = {k: [0, 0.0] for k in BK}
    for x in rs:
        k = bucket(x['diff']); b[k][0] += 1; b[k][1] += x['tot']
    ab = [x['ab'] for x in rs if isinstance(x['ab'], (int, float))]
    out['buckets'][m] = {'rows': b, 'items': len(rs), 'value': sum(x['tot'] for x in rs),
                         'leadAvg': sum(ab) / len(ab), 'leadMed': sorted(ab)[len(ab) // 2]}

lv, li, ld, lm = defaultdict(float), defaultdict(int), defaultdict(list), defaultdict(set)
for x in rows:
    if x['diff'] < 0:
        lv[x['sup']] += x['tot']; li[x['sup']] += 1; ld[x['sup']].append(x['diff']); lm[x['sup']].add(x['m'])
out['lateSup'] = [{'s': s, 'v': round(v, 2), 'n': li[s],
                   'avg': round(sum(ld[s]) / len(ld[s]), 1), 'worst': min(ld[s]),
                   'm': ','.join(sorted(lm[s], key=lambda z: M.index(z)))}
                  for s, v in sorted(lv.items(), key=lambda kv: -kv[1])]

# ---------------------------------------------------------------- 3) Supplier / Pending (ฐาน PO)
html = open('procurement_report.html', encoding='utf-8').read()
js = re.findall(r'<script[^>]*>(.*?)</script>', html, re.S)[1]
grab = lambda name: json.loads([l for l in js.split('\n') if l.startswith(f'const {name} =')][0]
                               .split('= ', 1)[1].rstrip(';'))
OI, MD = grab('ORDER_ITEMS'), grab('MD')

THAI = re.compile(r'[฀-๿]')
def origin(s):
    """ประมาณการในประเทศ/ต่างประเทศจากชื่อ — ต้องให้ทีมยืนยัน"""
    u = s.upper()
    if THAI.search(s) or 'THAILAND' in u or '(ไทย' in s:
        return 'ในประเทศ'
    return 'ต่างประเทศ'

sv, si, sgrp, smon = defaultdict(float), defaultdict(int), defaultdict(lambda: defaultdict(float)), defaultdict(set)
po_val, po_sup, po_grp, po_mon = defaultdict(float), {}, {}, {}
for m in M:
    for x in OI.get(m, []):
        s = (x['s'] or '').strip(); v = x.get('t') or 0
        sv[s] += v; si[s] += 1; sgrp[s][MD['groups'][x['g']]] += v; smon[s].add(m)
        p = x.get('po')
        if p:
            po_val[p] += v; po_sup[p] = s; po_grp[p] = MD['groups'][x['g']]; po_mon[p] = m

TOT = sum(sv.values())
ranked = sorted(sv.items(), key=lambda kv: -kv[1])
out['supTotal'] = round(TOT, 2)
out['supCount'] = len(sv)
out['topSup'] = [{'s': s, 'v': round(v, 2), 'n': si[s],
                  'g': max(sgrp[s].items(), key=lambda kv: kv[1])[0],
                  'm': ','.join(sorted(smon[s], key=lambda z: M.index(z))),
                  'o': origin(s)} for s, v in ranked[:20]]
vals = [v for _, v in ranked]
out['conc'] = {n: round(sum(vals[:n]) / TOT * 100, 1) for n in (1, 3, 5, 10, 20)}
out['pareto'] = [round(sum(vals[:i + 1]) / TOT * 100, 2) for i in range(min(30, len(vals)))]

# แยกในประเทศ / ต่างประเทศ
oo = defaultdict(float)
for s, v in sv.items():
    oo[origin(s)] += v
out['origin'] = {k: round(v, 2) for k, v in oo.items()}

recv_po = set(x['po'] for x in rows if x['po'])
out['pending'] = {}
for m in ['Jun', 'Jul']:
    pos = [p for p in po_val if po_mon[p] == m]
    miss = [p for p in pos if p not in recv_po]
    out['pending'][m] = {
        'poCount': len(pos), 'poValue': round(sum(po_val[p] for p in pos), 2),
        'missCount': len(miss), 'missValue': round(sum(po_val[p] for p in miss), 2),
        'top': [{'po': p, 'v': round(po_val[p], 2), 's': po_sup[p], 'g': po_grp[p]}
                for p in sorted(miss, key=lambda p: -po_val[p])[:10]]}

# ---------------------------------------------------------------- 4) Remark จากไฟล์ Cost Saving
out['csRemark'] = []
for m, f in [('Jun', 'data/Cost_Saving_Jun_2026.xlsx'), ('Jul', 'data/Cost_Saving_Jul_2026.xlsx')]:
    w = openpyxl.load_workbook(f, data_only=True); s = w[w.sheetnames[0]]
    for r in range(5, s.max_row + 1):
        det, rmk = s.cell(r, 2).value, s.cell(r, 10).value
        if det and rmk and str(det).strip() != "Total's":
            out['csRemark'].append({'m': m, 'd': str(det).strip(), 'r': str(rmk).strip(),
                                    'save': s.cell(r, 8).value})

json.dump(out, open('data/payload.json', 'w'), ensure_ascii=False, indent=1, default=str)

# ---------------------------------------------------------------- สรุปหน้าจอ
print('OTD Jan–Jul  :', ' '.join(f'{m} {p:.1f}%' for m, p in zip(M, out['otd']['pct'])))
print(f"OTD รวม      : {out['otd']['totPct']:.2f}%  ({out['otd']['totOn']:.0f}/{out['otd']['totPo']:.0f})  เป้า {out['otd']['target']:.0f}%")
print(f"Supplier     : {out['supCount']} ราย | Top1 {out['conc'][1]}% Top5 {out['conc'][5]}% Top10 {out['conc'][10]}%")
print(f"ในประเทศ/ตปท.: {out['origin']}")
for m in ['Jun', 'Jul']:
    p = out['pending'][m]
    print(f"Pending {m}   : {p['missCount']}/{p['poCount']} ใบ ฿{p['missValue']:,.0f} จาก ฿{p['poValue']:,.0f} ({p['missValue']/p['poValue']*100:.1f}%)")
print(f"ซัพส่งช้า     : {len(out['lateSup'])} ราย | Remark cost saving {len(out['csRemark'])} รายการ")
print('-> data/payload.json')
